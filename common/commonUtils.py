import hashlib
import io
import os
import random
import string
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from common.locales import t


def text2md5(text, project='starter'):
    hl = hashlib.md5()
    hl.update((text + f'_dr34m_cn_{project}').encode(encoding='utf-8'))
    return hl.hexdigest()


def passwd2md5(passwd):
    from common.config import getConfig
    cfg = getConfig()
    passwd_str = cfg['server']['key']
    return text2md5(passwd + passwd_str)


def stampToTime(stamp):
    """
    时间戳格式化为时间
    :param stamp: 时间戳，秒级
    :return: 格式化后的时间字符串，如2024-01-01 12:00:35
    """
    return datetime.fromtimestamp(stamp).strftime("%Y-%m-%d %H:%M:%S")


def timeToStamp(timeStr):
    """
    时间字符串转为时间戳
    :param timeStr: 时间字符串，如2024-01-01 12:00:35
    :return: 时间戳，秒级
    """
    return int(datetime.timestamp(datetime.strptime(timeStr, "%Y-%m-%d %H:%M:%S")))


def readOrSet(fileName, default, force=False):
    """
    从文件读取内容读取，不存在则创建
    :param fileName: 文件名，如data/111.txt
    :param default: 不存在文件时默认值，或强制覆盖值
    :param force: 强制用默认值覆盖
    :return: 结果
    """
    if os.path.exists(fileName) and force is False:
        with open(fileName, 'r') as file:
            fnData = file.read()
    else:
        fnData = default
        with open(fileName, 'w') as file:
            file.write(default)
    return fnData


def generatePasswd(length=8):
    """
    生成随机密码
    :param length: 密码长度
    :return:
    """
    characters = string.ascii_letters + string.digits
    return ''.join(random.choices(characters, k=length))


def onlySuperAdmin(req):
    """
    仅限超级管理员访问
    :param req:
    :param lang:
    :return:
    """
    userRole = req['__user']['role']
    if userRole != 0:
        raise Exception(t('permission.only_super_admin', req['__lang']))


def exportToExcel(tableData, widthMap=None):
    """
    导出到表格
    :param tableData: 字典列表
    :param widthMap: 各列宽度，如{"title": 20}
    :return:
    """
    wb = Workbook()
    ws = wb.active
    headers = list(tableData[0].keys())
    ws.append(headers)

    for row in tableData:
        ws.append([row.get(key, "") for key in headers])

    defaultWidth = 25
    if widthMap is None:
        widthMap = {}
    for i, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = widthMap.get(col_name, defaultWidth)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
